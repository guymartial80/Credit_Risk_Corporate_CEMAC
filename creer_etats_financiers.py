import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def verifier_longueurs(data):
    """Vérifie que tous les tableaux ont la même longueur"""
    longueurs = {key: len(value) for key, value in data.items()}
    if len(set(longueurs.values())) != 1:
        print(f"❌ Erreur: Les tableaux n'ont pas la même longueur: {longueurs}")
        return False
    return True

def creer_bilan():
    """Crée le fichier Excel pour le bilan"""
    data = {
        'compte': ['211', '213', '214', '215', '218', '261', '267', '271', '275', '281'],
        'libellé': ['Terrains', 'Constructions', 'Installations techniques', 'Autres immobilisations', 
                   'Avances et acomptes', 'Dépôts et cautionnements', 'Prêts', 'Stocks', 
                   'En-cours', 'Créances clients'],
        '2021': [150000, 300000, 80000, 50000, 20000, 15000, 25000, 120000, 45000, 180000],
        '2022': [145000, 320000, 85000, 52000, 22000, 16000, 27000, 125000, 48000, 190000],
        '2023': [140000, 350000, 90000, 55000, 25000, 18000, 30000, 130000, 50000, 200000]
    }
    
    if not verifier_longueurs(data):
        # Ajustement automatique si nécessaire
        min_length = min(len(data['compte']), len(data['libellé']), len(data['2021']), len(data['2022']), len(data['2023']))
        for key in data:
            data[key] = data[key][:min_length]
    
    df = pd.DataFrame(data)
    return df

def creer_compte_resultat():
    """Crée le fichier Excel pour le compte de résultat"""
    data = {
        'compte': ['701', '702', '703', '704', '705', '706'],
        'libellé': ['Ventes de produits finis', 'Ventes de produits intermédiaires', 'Ventes de produits résiduels',
                   'Travaux', 'Études', 'Prestations de services'],
        '2021': [500000, 120000, 25000, 80000, 45000, 60000],
        '2022': [550000, 130000, 28000, 85000, 48000, 65000],
        '2023': [600000, 140000, 30000, 90000, 50000, 70000]
    }
    
    if not verifier_longueurs(data):
        min_length = min(len(data['compte']), len(data['libellé']), len(data['2021']), len(data['2022']), len(data['2023']))
        for key in data:
            data[key] = data[key][:min_length]
    
    df = pd.DataFrame(data)
    return df

def creer_flux_tresorerie():
    """Crée le fichier Excel pour le tableau de flux de trésorerie"""
    data = {
        'compte': ['7011', '7012', '7013', '7014', '7015', '7016'],
        'libellé': ['Encaissements clients', 'Encaissements autres produits', 'Encaissements produits financiers',
                   'Encaissements produits exceptionnels', 'Subventions d\'investissement reçues',
                   'Encaissements cessions d\'immobilisations'],
        '2021': [480000, 115000, 22000, 75000, 40000, 55000],
        '2022': [530000, 125000, 25000, 80000, 45000, 60000],
        '2023': [580000, 135000, 28000, 85000, 48000, 65000]
    }
    
    if not verifier_longueurs(data):
        min_length = min(len(data['compte']), len(data['libellé']), len(data['2021']), len(data['2022']), len(data['2023']))
        for key in data:
            data[key] = data[key][:min_length]
    
    df = pd.DataFrame(data)
    return df

def formater_fichier_excel(nom_fichier, df, titre):
    """Formate le fichier Excel avec mise en page professionnelle"""
    try:
        with pd.ExcelWriter(nom_fichier, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=titre, index=False)
            
            # Accéder au workbook et worksheet
            workbook = writer.book
            worksheet = writer.sheets[titre]
            
            # Définir les styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Formater l'en-tête
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 10,  # compte
                'B': 40,  # libellé
                'C': 15,  # 2021
                'D': 15,  # 2022
                'E': 15   # 2023
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formater les nombres avec séparateurs de milliers
            for row in range(2, worksheet.max_row + 1):
                for col in ['C', 'D', 'E']:
                    cell = worksheet[f'{col}{row}']
                    try:
                        value = float(cell.value)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass
                        
        return True
    except Exception as e:
        print(f"❌ Erreur lors du formatage du fichier {nom_fichier}: {e}")
        return False

def generer_fichiers_excel():
    """Génère les trois fichiers Excel demandés"""
    try:
        print("🔄 Début de la génération des fichiers Excel...")
        
        # Créer les dataframes avec vérification
        print("📊 Création du bilan...")
        bilan_df = creer_bilan()
        print(f"   ✓ Bilan créé: {bilan_df.shape[0]} lignes, {bilan_df.shape[1]} colonnes")
        
        print("📊 Création du compte de résultat...")
        compte_resultat_df = creer_compte_resultat()
        print(f"   ✓ Compte de résultat créé: {compte_resultat_df.shape[0]} lignes, {compte_resultat_df.shape[1]} colonnes")
        
        print("📊 Création du tableau de flux de trésorerie...")
        flux_tresorerie_df = creer_flux_tresorerie()
        print(f"   ✓ Flux de trésorerie créé: {flux_tresorerie_df.shape[0]} lignes, {flux_tresorerie_df.shape[1]} colonnes")
        
        # Générer les fichiers Excel
        print("\n💾 Génération des fichiers Excel...")
        
        succes_bilan = formater_fichier_excel('bilan_entreprise_2021_2023.xlsx', bilan_df, 'Bilan')
        succes_cr = formater_fichier_excel('compte_resultat_entreprise_2021_2023.xlsx', compte_resultat_df, 'Compte de Resultat')
        succes_flux = formater_fichier_excel('flux_tresorerie_entreprise_2021_2023.xlsx', flux_tresorerie_df, 'Flux de Tresorerie')
        
        if succes_bilan and succes_cr and succes_flux:
            print("\n✅ Tous les fichiers Excel ont été générés avec succès:")
            print("   - bilan_entreprise_2021_2023.xlsx")
            print("   - compte_resultat_entreprise_2021_2023.xlsx")
            print("   - flux_tresorerie_entreprise_2021_2023.xlsx")
            
            # Afficher un aperçu des données
            print("\n📊 Aperçu des données générées:")
            print("\nBILAN (3 premières lignes):")
            print(bilan_df.head(3))
            print("\nCOMPTE DE RÉSULTAT (3 premières lignes):")
            print(compte_resultat_df.head(3))
            print("\nFLUX DE TRÉSORERIE (3 premières lignes):")
            print(flux_tresorerie_df.head(3))
            
        else:
            print("\n⚠️ Certains fichiers n'ont pas pu être générés correctement")
            
    except Exception as e:
        print(f"❌ Erreur lors de la génération des fichiers: {e}")
        import traceback
        print(f"🔍 Détails de l'erreur:")
        print(traceback.format_exc())

# Test de vérification des données avant exécution
def tester_donnees():
    """Teste la cohérence des données avant génération"""
    print("🧪 Test de cohérence des données...")
    
    # Test bilan
    data_bilan = {
        'compte': ['211', '213', '214', '215', '218', '261', '267', '271', '275', '281'],
        'libellé': ['Terrains', 'Constructions', 'Installations techniques', 'Autres immobilisations', 
                   'Avances et acomptes', 'Dépôts et cautionnements', 'Prêts', 'Stocks', 
                   'En-cours', 'Créances clients'],
        '2021': [150000, 300000, 80000, 50000, 20000, 15000, 25000, 120000, 45000, 180000],
        '2022': [145000, 320000, 85000, 52000, 22000, 16000, 27000, 125000, 48000, 190000],
        '2023': [140000, 350000, 90000, 55000, 25000, 18000, 30000, 130000, 50000, 200000]
    }
    
    for key, value in data_bilan.items():
        print(f"   Bilan - {key}: {len(value)} éléments")
    
    # Vérification
    longueurs_bilan = [len(data_bilan[key]) for key in data_bilan]
    if len(set(longueurs_bilan)) == 1:
        print("   ✓ Bilan: OK")
    else:
        print("   ❌ Bilan: Incohérence détectée")

# Exécuter le programme
if __name__ == "__main__":
    tester_donnees()
    print("\n" + "="*50)
    generer_fichiers_excel()